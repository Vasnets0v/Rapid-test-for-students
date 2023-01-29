import random
import time
from openpyxl import Workbook
from __init__ import get_db


class ExelSheet:
    def __init__(self, topic, sql_request):
        self.file_name = str(topic) + ' ' + str(time.strftime("%d.%m.%Y"))
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = '1'
        self.raw_users_score = sql_request.execute(f"SELECT * FROM score_for_theme_{topic} order by id desc")
        self.user_score = []

        self.ws['A1'] = "№"
        self.ws['B1'] = "Прізвище та Імя"
        self.ws['C1'] = "Група"
        self.ws['D1'] = "Результат"
        self.ws['E1'] = "Завершено тест"

        for user in self.raw_users_score:
            self.user_score.append(user)

    def insert_user_data(self):

        for data in self.user_score:
            self.ws.append(data)

        self.wb.save(self.file_name + ".xlsx")


def get_all_records_from_table(topic):
    db = get_db()
    raw_content = db.cursor().execute(f"SELECT * FROM {topic}")

    content = [item for item in raw_content]

    return content


def get_all_topics():
    db = get_db()
    raw_topics = db.cursor().execute(f"SELECT topic_title FROM tests_info")

    topics = [topic[0] for topic in raw_topics]

    return topics


def get_all_info_about_topics():
    db = get_db()
    raw_records = db.cursor().execute(f"SELECT * FROM tests_info")

    content = []
    # The variable title contains place(index) where the name of the topic is stored in the array
    title = 1

    for raw_row in raw_records:
        row_records = []

        for column in raw_row:
            row_records.append(column)

        content.append(row_records)

    for row in range(len(content)):
        content[row][title] = get_clear_topic_name(content[row][title])

    return content


def get_clear_topic_name(topic):
    word_array = topic.split('_')
    clean_topic = ' '.join(map(str, word_array))
    
    return clean_topic


def get_array_clear_topics(topics):
    content = []

    for topic in topics:
        content.append(get_clear_topic_name(topic))

    return content


def conver_topic_id_into_title(id):
    db = get_db()
    topic = db.cursor().execute(f'SELECT topic_title FROM tests_info WHERE id = {id}').fetchone()[0]
    
    return topic


def get_question_from_db(num_of_questions, topic_id):
    topic = conver_topic_id_into_title(topic_id)

    db = get_db()
    sql_request = db.cursor()
    total_columns = sql_request.execute(f"SELECT count(*) FROM {topic}").fetchone()[0] + 1

    random_ids = random.sample(range(1, total_columns), num_of_questions)
    content = {}

    for i in range(num_of_questions):
        sql_request.execute(f"""SELECT question, img_question, answer_1, img_1, answer_2, img_2, answer_3, img_3, 
        answer_4, img_4, answer_5, img_5, answer_6, img_6 FROM {topic} WHERE id = '{random_ids[i]}'""")

        tuple_content = sql_request.fetchone()
        array_content = []

        for item in tuple_content:
            array_content.append(item)

        content[random_ids[i]] = array_content


    return content


def get_num_of_records_in_table(topic):
    db = get_db()
    num_of_records = db.cursor().execute(f"SELECT COUNT('id') FROM {topic}").fetchall()[0][0]

    return num_of_records


def get_mixed_order(answers):
    return random.sample(range(1, answers + 1), answers)


def get_new_img_name(img):
    current_time = str(time.time())
    raw_img_name = current_time.split('.')
    old_img_name = img.split('.')
    name = raw_img_name[0] + old_img_name[0] + raw_img_name[1] + str(random.randint(1, 9999)) + "." + old_img_name[1]

    return name


def process_user_result(correct_answers, user_answers):
    num_of_tests = len(correct_answers)
    score = 0

    for correct_ans, user_ans in zip(correct_answers, user_answers):
        if correct_answers[correct_ans] == user_answers[user_ans]:
            score += 1

    percentage_score = round((score * 100) / num_of_tests)

    return percentage_score
