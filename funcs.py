import random
import time
from openpyxl import Workbook
from app import sql_request


class ExelSheet:
    def __init__(self, topic):
        self.file_name = str(topic) + ' ' + str(time.strftime("%d.%m.%Y"))
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = '1'
        self.raw_users_score = sql_request.execute(f"SELECT * FROM score_for_theme_{topic} order by id desc")
        self.user_score = []

        self.ws['A1'] = "№"
        self.ws['B1'] = "Прізвище та Імя"
        self.ws['C1'] = "Група"
        self.ws['D1'] = "Результат"
        self.ws['E1'] = "Завершено тест"

        for user in self.raw_users_score:
            self.user_score.append(user)

    def insert_user_data(self):

        for data in self.user_score:
            self.ws.append(data)

        self.wb.save(self.file_name + ".xlsx")


def get_all_records_from_table(topic):
    raw_content = sql_request.execute(f"SELECT * FROM {topic}")
    content = []

    for item in raw_content:
        content.append(item)

    return content


def get_num_of_records_in_table(topic):
    num_of_records = sql_request.execute(f"SELECT COUNT('id') FROM {topic}")

    return num_of_records.fetchall()[0][0]


def get_mixed_order(answers):
    return random.sample(range(1, answers + 1), answers)


def get_new_img_name(img):
    current_time = str(time.time())
    raw_img_name = current_time.split('.')
    old_img_name = img.split('.')
    name = raw_img_name[0] + old_img_name[0] + raw_img_name[1] + str(random.randint(1, 9999)) + "." + old_img_name[1]

    return name


def process_user_result(correct_answers, user_answers):
    num_of_tests = len(correct_answers)
    score = 0

    for correct_ans, user_ans in zip(correct_answers, user_answers):
        if correct_answers[correct_ans] == user_answers[user_ans]:
            score += 1

    percentage_score = round((score * 100) / num_of_tests)

    return percentage_score


def get_all_tables_from_db():
    all_table_content = sql_request.execute("SELECT * FROM sqlite_master WHERE type='table'")
    table_list = []

    for i in all_table_content:
        table_list.append(i[1])

    return table_list


def get_all_topics():
    raw_topics = sql_request.execute(f"SELECT topic_title FROM tests_info")
    topics = []

    for topic in raw_topics:
        topics.append(topic[0])

    return topics


def get_all_info_about_topics():
    raw_content = sql_request.execute(f"SELECT * FROM tests_info")

    content = []

    for item in raw_content:
        content.append(item)

    return content


def get_content_from_db(num_of_questions, topic):
    sql_request.execute(f"SELECT count(*) FROM {topic}")
    total_columns = sql_request.fetchone()[0] + 1

    random_ids = random.sample(range(1, total_columns), num_of_questions)
    content = {}

    for i in range(num_of_questions):
        sql_request.execute(f"""SELECT question, img_question, answer_1, img_1, answer_2, img_2, answer_3, img_3, 
        answer_4, img_4, answer_5, img_5, answer_6, img_6 FROM {topic} WHERE id = '{random_ids[i]}'""")

        tuple_content = sql_request.fetchone()
        array_content = []

        for item in tuple_content:
            array_content.append(item)

        content[random_ids[i]] = array_content

    return content
